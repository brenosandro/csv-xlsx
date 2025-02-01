# %%
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import os  # Import necessário para manipular caminhos de arquivos
from openpyxl import load_workbook

# Função para selecionar o arquivo CSV
def selecionar_csv():
    caminho = filedialog.askopenfilename(
        title="Selecione o arquivo CSV",
        filetypes=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")]
    )
    entrada_var.set(caminho)

# Função para processar o arquivo
def processar_arquivo():
    entrada = entrada_var.get()

    if not entrada:
        messagebox.showerror("Erro", "Por favor, selecione o arquivo de entrada.")
        return

    try:
        # Gera o caminho de saída no mesmo diretório e com o mesmo nome do arquivo de entrada
        diretorio, nome_arquivo = os.path.split(entrada)
        nome_base, _ = os.path.splitext(nome_arquivo)
        saida = os.path.join(diretorio, f"{nome_base}.xlsx")

        # Lê o arquivo CSV
        df = pd.read_csv(entrada, delimiter=',')  # Delimitador ',' já é padrão

        # Verifica e ajusta o tipo de dados em cada coluna
        for coluna in df.columns:
            try:
                # Tenta converter para números (float)
                df[coluna] = pd.to_numeric(df[coluna], errors='raise')
            except ValueError:
                # Mantém como texto se não for número
                df[coluna] = df[coluna].astype(str)

        # Calcula o subtotal das colunas numéricas
        subtotais = df.select_dtypes(include='number').sum()

        # Salva o DataFrame no Excel sem incluir a linha de subtotal
        df.to_excel(saida, index=False, engine='openpyxl')

        # Adiciona a linha de subtotal acima do cabeçalho no arquivo Excel
        workbook = load_workbook(saida)
        sheet = workbook.active

        # Insere a linha de subtotal na primeira linha
        sheet.insert_rows(1)

        # Insere os valores do subtotal na respectiva coluna
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in subtotais.index:  # Verifica se a coluna é numérica
                sheet.cell(row=1, column=col_idx, value=subtotais[col_name])
            else:
                sheet.cell(row=1, column=col_idx, value="Subtotal")  # Rótulo para as colunas não numéricas

        # Salva o arquivo Excel com a linha de subtotal
        workbook.save(saida)
        messagebox.showinfo("Sucesso", f"Arquivo processado e salvo em: {saida}")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

# Cria a interface gráfica
janela = tk.Tk()
janela.title("Processar CSV para Excel")

# Variável para armazenar o caminho do arquivo de entrada
entrada_var = tk.StringVar()

# Elementos da interface
tk.Label(janela, text="Selecione o arquivo CSV:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
tk.Entry(janela, textvariable=entrada_var, width=50).grid(row=0, column=1, padx=10, pady=10)
tk.Button(janela, text="Procurar", command=selecionar_csv).grid(row=0, column=2, padx=10, pady=10)

tk.Button(janela, text="Processar", command=processar_arquivo).grid(row=1, column=0, columnspan=3, pady=20)

# Executa a interface
janela.mainloop()
